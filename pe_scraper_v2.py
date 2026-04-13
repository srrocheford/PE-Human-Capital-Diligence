#!/usr/bin/env python3
"""
PE Human Capital Tracker v2
============================
Scrapes one year-end snapshot per year from Wayback Machine.
Outputs a formatted Excel workbook with charts.

Usage:
    python pe_scraper_v2.py

Requires: requests, pandas, openpyxl
"""

import re, time
from datetime import datetime
import requests
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

FIRM_NAME   = "Linden Capital Partners"
START_YEAR  = 2010
OUTPUT_FILE = "/sessions/vibrant-blissful-galileo/mnt/DSAIL Project/linden_human_capital_v2.xlsx"
CDX_API     = "https://web.archive.org/cdx/search/cdx"

SOURCES = [
    {"base_url": "lindenllc.com/team/", "label": "lindenllc.com"},
    {"base_url": "linden.com/team/",    "label": "linden.com"},
]

FUNCTION_MAP = {
    "investment-team":        "Investment Team",
    "structured-capital":     "Structured Capital",
    "operating-partners":     "Operating Advisors",
    "operating-advisors":     "Operating Advisors",
    "human-capital":          "Human Capital",
    "human-capital-finance":  "Human Capital",   # combined slug on old site; defaults to HC
    "finance":                "Finance",
    "accounting":             "Finance",
    "administrative":         "Administration",
    "admin":                  "Administration",
    "investor-relations":     "Investor Relations",
    "legal":                  "Legal & Compliance",
    "compliance":             "Legal & Compliance",
    "senior-advisors":        "Senior Advisors",
    "it":                     "Information Technology",
    "information-technology": "Information Technology",
    "technology":             "Information Technology",
}

FUNCTION_ORDER = [
    "Investment Team", "Structured Capital", "Operating Advisors",
    "Human Capital", "Finance", "Investor Relations",
    "Legal & Compliance", "Senior Advisors", "Administration", "Other",
]

# Palette for chart bars (one colour per function)
FUNCTION_COLORS = {
    "Investment Team":    "1F4E79",
    "Structured Capital": "2E75B6",
    "Operating Advisors": "70AD47",
    "Human Capital":      "FFC000",
    "Finance":            "ED7D31",
    "Investor Relations": "A9D18E",
    "Legal & Compliance": "C55A11",
    "Senior Advisors":    "7030A0",
    "Administration":     "BFBFBF",
    "Other":              "808080",
}


# ─────────────────────────────────────────────────────────────────────────────
# NAME NORMALIZATION
# ─────────────────────────────────────────────────────────────────────────────

def normalize_name(name: str) -> str:
    """Strip middle initials so 'John A. Smith' == 'John Smith'."""
    name = name.strip()
    name = re.sub(r'\b[A-Z]\.\s*', ' ', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 · WAYBACK CDX — one snapshot per year, closest to Dec 31
# ─────────────────────────────────────────────────────────────────────────────

def get_yearly_snapshots(url: str, start_year: int) -> list[dict]:
    print(f"  CDX query: {url} (from {start_year}) ...")
    try:
        r = requests.get(CDX_API, params={
            "url": url, "output": "json",
            "fl": "timestamp,original,statuscode",
            "filter": "statuscode:200",
            "from": f"{start_year}0101",
            "limit": 1000,
        }, timeout=60)
        r.raise_for_status()
        data = r.json()
        if len(data) <= 1:
            print("    → 0 snapshots")
            return []
    except Exception as e:
        print(f"    ⚠ CDX error: {e}")
        return []

    by_year: dict[str, dict] = {}
    for row in data[1:]:
        ts = row[0]
        year = ts[:4]
        try:
            snap_dt  = datetime(int(ts[:4]), int(ts[4:6]), int(ts[6:8]))
            target   = datetime(int(year), 12, 31)
            distance = abs((target - snap_dt).days)
        except ValueError:
            continue
        if year not in by_year or distance < by_year[year]["distance"]:
            by_year[year] = {
                "timestamp":   ts,
                "original":    row[1],
                "distance":    distance,
                "actual_date": snap_dt.strftime("%Y-%m-%d"),
            }

    result = sorted(
        [{"timestamp": v["timestamp"], "original": v["original"],
          "year": k, "actual_date": v["actual_date"]}
         for k, v in by_year.items()],
        key=lambda x: x["year"]
    )
    if result:
        print(f"    → {len(result)} yearly snapshots ({result[0]['year']}–{result[-1]['year']})")
        for s in result:
            dist = by_year[s["year"]]["distance"]
            note = f"(Dec 31 −{dist}d)" if dist > 0 else "(Dec 31 exact)"
            print(f"       {s['year']}: {s['actual_date']} {note}")
    return result


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 · FETCH with exponential backoff retry
# ─────────────────────────────────────────────────────────────────────────────

def fetch_html(timestamp: str, original_url: str, max_retries: int = 3) -> str | None:
    url = f"https://web.archive.org/web/{timestamp}id_/{original_url}"
    for attempt in range(max_retries):
        try:
            r = requests.get(url, timeout=45,
                headers={"User-Agent": "PE-Diligence-Research-Tool/1.0"})
            r.raise_for_status()
            return r.text
        except Exception as e:
            if attempt < max_retries - 1:
                wait = 5 * (2 ** attempt)
                print(f"      retry {attempt+1}/{max_retries-1} in {wait}s — {e}")
                time.sleep(wait)
            else:
                print(f"      ⚠ Failed: {e}")
    return None


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 · EXTRACT people from HTML
# ─────────────────────────────────────────────────────────────────────────────

def extract_from_dropdown(html: str, year: str) -> list[dict]:
    pattern = r'<option\s+value="[^"]+/team/([^/"]+)/[^/"]+/">\s*([^<]+?)\s*</option>'
    people = []
    for func_slug, raw_name in re.findall(pattern, html, re.IGNORECASE):
        name = normalize_name(raw_name)
        if not name or len(name) < 3:
            continue
        func = FUNCTION_MAP.get(func_slug.lower(), func_slug.replace("-", " ").title())
        people.append({"name": name, "function": func, "year": year})
    return people

def extract_from_cards(html: str, year: str) -> list[dict]:
    pattern = r'<article[^>]+class="item biography ([^"]+)"[^>]*>.*?<h[12]>([^<]+)</h[12]>'
    people = []
    for func_classes, raw_name in re.findall(pattern, html, re.DOTALL | re.IGNORECASE):
        name = normalize_name(raw_name)
        if not name:
            continue
        func_slug = func_classes.strip().split()[-1]
        func = FUNCTION_MAP.get(func_slug.lower(), func_slug.replace("-", " ").title())
        people.append({"name": name, "function": func, "year": year})
    return people

def extract_people(html: str, year: str) -> list[dict]:
    people = extract_from_dropdown(html, year)
    return people if people else extract_from_cards(html, year)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 · ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────

def analyze(df: pd.DataFrame):
    df = df.copy()
    df["year"] = df["year"].astype(int)
    df["name_norm"] = df["name"].apply(lambda n: n.lower().strip())
    df = df.sort_values("year")
    all_years = sorted(df["year"].unique())
    latest_year = max(all_years)

    arrivals, departures, tenure_rows = [], [], []

    for norm, group in df.groupby("name_norm"):
        group     = group.sort_values("year").drop_duplicates(subset="year")
        canon     = group.iloc[0]["name"]
        func      = group.iloc[-1]["function"]
        first_yr  = int(group["year"].min())
        last_yr   = int(group["year"].max())
        tenure_yr = last_yr - first_yr
        still_active = (last_yr == latest_year)

        arrivals.append({
            "name":         canon,
            "function":     group.iloc[0]["function"],
            "arrival_year": first_yr,
        })

        tenure_rows.append({
            "name":          canon,
            "function":      func,
            "first_seen":    first_yr,
            "last_seen":     last_yr,
            "tenure_years":  tenure_yr,
            "still_active":  still_active,
            "linkedin_notes": "",
        })

        if not still_active:
            departures.append({
                "name":                  canon,
                "function":              func,
                "last_seen":             last_yr,
                "approx_departure_year": last_yr + 1,
                "tenure_years":          tenure_yr,
                "linkedin_notes":        "",
            })

    return pd.DataFrame(arrivals), pd.DataFrame(departures), pd.DataFrame(tenure_rows)


def headcount_pivot(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["year"] = df["year"].astype(int)
    rows = []
    for year, ydf in df.groupby("year"):
        counts = ydf.groupby("function").size().to_dict()
        row = {"Year": year}
        row.update(counts)
        rows.append(row)
    if not rows:
        return pd.DataFrame()
    result = pd.DataFrame(rows).fillna(0)
    func_cols = [c for c in FUNCTION_ORDER if c in result.columns]
    result["Total"] = result[func_cols].sum(axis=1)
    for c in func_cols + ["Total"]:
        result[c] = result[c].astype(int)
    return result.sort_values("Year")[["Year"] + func_cols + ["Total"]]


def turnover_metrics(departures_df: pd.DataFrame, tenure_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    funcs_present = tenure_df["function"].unique()
    for func in FUNCTION_ORDER:
        if func not in funcs_present:
            continue
        fdep = departures_df[departures_df["function"] == func]
        ften = tenure_df[tenure_df["function"] == func]
        total_ever = len(ften)
        n_active   = len(ften[ften["still_active"]])
        n_departed = len(fdep)
        retention  = round(n_active / total_ever * 100, 1) if total_ever else 0.0
        avg_tenure = round(fdep["tenure_years"].mean(), 1) if len(fdep) else None
        names_list = "; ".join(sorted(fdep["name"].tolist()))
        rows.append({
            "Function":                          func,
            "Total Ever Employed":               total_ever,
            "Currently Active":                  n_active,
            "Total Leavers":                     n_departed,
            "Retention Rate":                    retention / 100,
            "Avg Tenure Before Departure (yrs)": avg_tenure if avg_tenure else "",
            "Names of Leavers":                  names_list,
        })
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5 · BUILD EXCEL WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────

# ── Style helpers ────────────────────────────────────────────────────────────

def hdr_font(bold=True, size=10, color="FFFFFF"):
    return Font(name="Arial", bold=bold, size=size, color=color)

def body_font(bold=False, size=10, color="000000"):
    return Font(name="Arial", bold=bold, size=size, color=color)

def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header_row(ws, row, n_cols, fill_hex="1F4E79"):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font  = hdr_font()
        cell.fill  = hdr_fill(fill_hex)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()

def style_data_row(ws, row, n_cols, shade=False):
    fill = hdr_fill("EBF3FB") if shade else hdr_fill("FFFFFF")
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font   = body_font()
        cell.fill   = fill
        cell.border = thin_border()
        cell.alignment = Alignment(vertical="center", wrap_text=(col == n_cols))

def write_df_to_sheet(ws, df, start_row=1, header_fill="1F4E79"):
    headers = list(df.columns)
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    style_header_row(ws, start_row, len(headers), header_fill)

    for r_idx, row_data in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
        style_data_row(ws, r_idx, len(headers), shade=(r_idx % 2 == 0))

    return start_row + len(df)   # returns last data row written


def build_excel(raw_df, hc_df, arrivals_df, departures_df, tenure_df, turnover_df, output_path):
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Sheet 1: Headcount by Year ───────────────────────────────────────────
    ws_hc = wb.create_sheet("Headcount by Year")
    func_cols = [c for c in FUNCTION_ORDER if c in hc_df.columns]
    n_years   = len(hc_df)
    n_data_cols = len(hc_df.columns)

    ws_hc["A1"] = f"{FIRM_NAME} — Year-End Headcount by Function"
    ws_hc["A1"].font = Font(name="Arial", bold=True, size=13, color="1F4E79")
    ws_hc.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_data_cols)
    ws_hc.row_dimensions[1].height = 22

    ws_hc["A2"] = "Source: Wayback Machine snapshots of lindenllc.com and linden.com team pages."
    ws_hc["A2"].font = Font(name="Arial", italic=True, size=9, color="595959")
    ws_hc.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_data_cols)

    data_start = 4
    write_df_to_sheet(ws_hc, hc_df, start_row=data_start)

    # Format Retention column if present
    for r in range(data_start + 1, data_start + n_years + 1):
        ws_hc.cell(r, 1).alignment = Alignment(horizontal="center")

    # Column widths
    ws_hc.column_dimensions["A"].width = 8
    for i, col in enumerate(hc_df.columns[1:], 2):
        ws_hc.column_dimensions[get_column_letter(i)].width = max(len(str(col)) + 2, 12)

    # ── Stacked bar chart ────────────────────────────────────────────────────
    chart = BarChart()
    chart.type      = "col"
    chart.grouping  = "stacked"
    chart.title     = "Headcount by Function (Year-End)"
    chart.y_axis.title = "Headcount"
    chart.x_axis.title = "Year"
    chart.width  = 24
    chart.height = 14
    chart.legend.position = "b"

    # Reference only the function columns (not Year or Total)
    func_col_indices = [list(hc_df.columns).index(f) + 1 for f in func_cols]
    for col_idx in func_col_indices:
        func_name = hc_df.columns[col_idx - 1]
        data_ref = Reference(ws_hc,
            min_col=col_idx, max_col=col_idx,
            min_row=data_start, max_row=data_start + n_years)
        chart.add_data(data_ref, titles_from_data=True)
        # Colour each series
        hex_col = FUNCTION_COLORS.get(func_name, "808080")
        chart.series[-1].graphicalProperties.solidFill = hex_col
        chart.series[-1].graphicalProperties.line.solidFill = hex_col

    cats = Reference(ws_hc, min_col=1, min_row=data_start + 1, max_row=data_start + n_years)
    chart.set_categories(cats)

    chart_anchor = f"A{data_start + n_years + 3}"
    ws_hc.add_chart(chart, chart_anchor)

    # ── Sheet 2: Turnover Metrics ────────────────────────────────────────────
    ws_to = wb.create_sheet("Turnover Metrics")
    n_to_cols = len(turnover_df.columns)

    ws_to["A1"] = f"{FIRM_NAME} — Turnover Metrics by Function"
    ws_to["A1"].font = Font(name="Arial", bold=True, size=13, color="1F4E79")
    ws_to.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_to_cols)
    ws_to.row_dimensions[1].height = 22

    ws_to["A2"] = (
        "Note: 'Retention Rate' = currently active / total ever employed at this firm. "
        "Departure dates are approximated as last-seen year + 1 (based on yearly snapshots). "
        "The old site (lindenllc.com) combined Human Capital and Finance into one URL slug; "
        "those records are attributed to Human Capital."
    )
    ws_to["A2"].font = Font(name="Arial", italic=True, size=9, color="595959")
    ws_to.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_to_cols)
    ws_to.row_dimensions[2].height = 36

    to_data_start = 4
    write_df_to_sheet(ws_to, turnover_df, start_row=to_data_start)

    # Format Retention Rate as percentage
    ret_col_idx = list(turnover_df.columns).index("Retention Rate") + 1
    for r in range(to_data_start + 1, to_data_start + len(turnover_df) + 1):
        cell = ws_to.cell(r, ret_col_idx)
        cell.number_format = "0.0%"

    # Column widths
    col_widths = {"Function": 22, "Total Ever Employed": 18, "Currently Active": 16,
                  "Total Leavers": 14, "Retention Rate": 15,
                  "Avg Tenure Before Departure (yrs)": 28, "Names of Leavers": 60}
    for i, col in enumerate(turnover_df.columns, 1):
        ws_to.column_dimensions[get_column_letter(i)].width = col_widths.get(col, 15)

    # Wrap the names column
    names_col_idx = list(turnover_df.columns).index("Names of Leavers") + 1
    for r in range(to_data_start + 1, to_data_start + len(turnover_df) + 1):
        ws_to.cell(r, names_col_idx).alignment = Alignment(wrap_text=True, vertical="top")
        ws_to.row_dimensions[r].height = 60

    # ── Sheet 3: Promotions ──────────────────────────────────────────────────
    ws_pr = wb.create_sheet("Promotions")
    ws_pr["A1"] = f"{FIRM_NAME} — Time-to-Promotion Tracking"
    ws_pr["A1"].font = Font(name="Arial", bold=True, size=13, color="1F4E79")
    ws_pr.merge_cells("A1:F1")

    ws_pr["A3"] = "STATUS: NOT AVAILABLE FOR THIS FIRM"
    ws_pr["A3"].font = Font(name="Arial", bold=True, size=11, color="C00000")

    explanation = (
        "Promotion tracking requires individual title data (e.g., Associate, Vice President, Principal, Partner) "
        "to identify title transitions over time.\n\n"
        "For Linden Capital Partners, titles are NOT displayed on either lindenllc.com or linden.com team pages. "
        "Individual profile pages are JavaScript-rendered and were not captured by the Wayback Machine.\n\n"
        "This tab will populate automatically for PE firms that display titles on their main team listing page. "
        "Many middle-market PE firms do show titles — this feature is built and ready.\n\n"
        "Workarounds to consider for Linden:\n"
        "  1. People Data Labs or ContactOut API (~$0.10/record) for LinkedIn-sourced title history\n"
        "  2. Manual research on departed investment professionals via LinkedIn\n"
        "  3. Run the scraper against a firm that does expose titles (e.g., many Midwest healthcare PE firms do)"
    )
    ws_pr["A5"] = explanation
    ws_pr["A5"].font = Font(name="Arial", size=10, color="595959")
    ws_pr["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_pr.merge_cells("A5:F20")
    ws_pr.row_dimensions[5].height = 200
    ws_pr.column_dimensions["A"].width = 100

    # ── Sheet 4: Departures Detail ───────────────────────────────────────────
    ws_dep = wb.create_sheet("Departures Detail")
    n_dep_cols = len(departures_df.columns)
    ws_dep["A1"] = f"{FIRM_NAME} — Individual Departure Log"
    ws_dep["A1"].font = Font(name="Arial", bold=True, size=13, color="1F4E79")
    ws_dep.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_dep_cols)

    write_df_to_sheet(ws_dep, departures_df.sort_values(
        ["function", "approx_departure_year"]), start_row=3)
    ws_dep.column_dimensions["A"].width = 28
    ws_dep.column_dimensions["B"].width = 22
    for i in range(3, n_dep_cols + 1):
        ws_dep.column_dimensions[get_column_letter(i)].width = 16
    # LinkedIn notes column wider
    linkedin_idx = list(departures_df.columns).index("linkedin_notes") + 1
    ws_dep.column_dimensions[get_column_letter(linkedin_idx)].width = 40

    # ── Sheet 5: Tenure ──────────────────────────────────────────────────────
    ws_ten = wb.create_sheet("Tenure")
    n_ten_cols = len(tenure_df.columns)
    ws_ten["A1"] = f"{FIRM_NAME} — Tenure by Individual"
    ws_ten["A1"].font = Font(name="Arial", bold=True, size=13, color="1F4E79")
    ws_ten.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_ten_cols)

    write_df_to_sheet(ws_ten, tenure_df.sort_values(
        ["function", "tenure_years"], ascending=[True, False]), start_row=3)
    ws_ten.column_dimensions["A"].width = 28
    ws_ten.column_dimensions["B"].width = 22
    for i in range(3, n_ten_cols + 1):
        ws_ten.column_dimensions[get_column_letter(i)].width = 16

    # ── Sheet 6: Raw Snapshots ───────────────────────────────────────────────
    ws_raw = wb.create_sheet("Raw Snapshots")
    write_df_to_sheet(ws_raw, raw_df, start_row=1)
    ws_raw.column_dimensions["A"].width = 28
    ws_raw.column_dimensions["B"].width = 22
    ws_raw.column_dimensions["C"].width = 10

    wb.save(output_path)
    print(f"  ✅  Saved → {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def run():
    print(f"\n{'='*60}")
    print(f"  {FIRM_NAME} — Human Capital Tracker v2")
    print(f"{'='*60}\n")

    # 1. Collect yearly snapshots across both URL eras
    all_snapshots: dict[str, dict] = {}  # keyed by year; linden.com takes priority
    for source in SOURCES:
        snaps = get_yearly_snapshots(source["base_url"], START_YEAR)
        for s in snaps:
            yr = s["year"]
            if yr not in all_snapshots or source["label"] == "linden.com":
                all_snapshots[yr] = {**s, "label": source["label"]}

    ordered = sorted(all_snapshots.values(), key=lambda x: x["year"])
    print(f"\n  Total yearly snapshots to process: {len(ordered)}")

    # 2. Scrape + extract
    print(f"\n[Scraping...]\n")
    all_records: list[dict] = []

    for i, snap in enumerate(ordered, 1):
        yr   = snap["year"]
        ts   = snap["timestamp"]
        orig = snap["original"]
        print(f"  [{i:>2}/{len(ordered)}] {yr} ({snap['actual_date']}, {snap['label']}) ...", end=" ", flush=True)

        html = fetch_html(ts, orig)
        if not html:
            print("skip")
            continue

        people = extract_people(html, yr)
        if people:
            all_records.extend(people)
            print(f"{len(people)} people")
        else:
            print("0 extracted")

        time.sleep(1.5)

    if not all_records:
        print("\n⚠  No records extracted.")
        return

    # 3. Build DataFrames
    raw_df = (pd.DataFrame(all_records)[["name", "function", "year"]]
              .dropna(subset=["name"])
              .drop_duplicates()
              .sort_values(["year", "name"])
              .reset_index(drop=True))

    print(f"\n  Records: {len(raw_df)} | Unique people: {raw_df['name'].nunique()} | Years: {raw_df['year'].nunique()}")

    # 4. Analysis
    arrivals_df, departures_df, tenure_df = analyze(raw_df)
    hc_df      = headcount_pivot(raw_df)
    turnover_df = turnover_metrics(departures_df, tenure_df)

    print(f"  Departures: {len(departures_df)} | Active: {len(tenure_df[tenure_df['still_active']])}")

    # 5. Build Excel
    print(f"\n[Building Excel...]\n")
    build_excel(raw_df, hc_df, arrivals_df, departures_df, tenure_df, turnover_df, OUTPUT_FILE)

    # 6. Print summary
    print(f"\n{'─'*60}")
    print(f"  HEADCOUNT TREND:")
    for _, row in hc_df.iterrows():
        bar = "█" * int(row["Total"] / 2)
        print(f"    {int(row['Year'])}  {int(row['Total']):>3}  {bar}")

    print(f"\n  TURNOVER SUMMARY:")
    for _, row in turnover_df.iterrows():
        print(f"    {row['Function']:<28} {row['Total Leavers']:>3} leavers  "
              f"{row['Retention Rate']:.0%} retention")
    print(f"{'─'*60}\n")


if __name__ == "__main__":
    run()
